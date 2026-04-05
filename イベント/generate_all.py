#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
96件のイベントフォルダ・レポート・プレゼンPDF・Webサイトを一括生成
"""

import os
import re
import openpyxl
from datetime import datetime

# ===== PDF生成用 =====
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Frame, PageTemplate, BaseDocTemplate
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# 日本語フォント登録
pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))

FONT_MINCHO = 'HeiseiMin-W3'
FONT_GOTHIC = 'HeiseiKakuGo-W5'

BASE_DIR = '/Users/ishiharajunichi/Desktop/北総フォルダ２/イベント'
EXCEL_PATH = os.path.join(BASE_DIR, '成田富里地域_2025-2026年度イベントスケジュール20250916.xlsx')

# ===== Excelからイベント読み込み =====
def load_events():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['【千葉メディア提出用】日付順_2026年度年間スケジュール']
    events = []
    idx = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[1] and hasattr(row[1], 'year'):
            idx += 1
            date_str = row[1].strftime('%Y-%m-%d')
            time_val = row[2]
            if hasattr(time_val, 'strftime'):
                time_str = time_val.strftime('%H:%M')
            elif time_val:
                time_str = str(time_val)[:5]
            else:
                time_str = ''
            title = str(row[5]).strip() if row[5] else ''
            # フォルダ名用にタイトルを短縮・クリーン
            clean_title = re.sub(r'[\/\\:*?"<>|]', '', title)[:40]
            events.append({
                'idx': idx,
                'date': date_str,
                'time': time_str,
                'venue': str(row[3]).strip() if row[3] else '',
                'circle': str(row[4]).strip() if row[4] else '',
                'title': title,
                'member': str(row[6]).strip() if row[6] else '',
                'folder_name': f"{date_str}_{clean_title}_{str(row[3]).strip() if row[3] else ''}",
            })
    return events

# ===== サークル別の詳細コンテンツ生成 =====
def generate_report_content(ev):
    """イベントタイトルとサークル名からレポート内容を生成"""
    circle = ev['circle']
    title = ev['title']
    venue = ev['venue']
    date = ev['date']
    time = ev['time']
    member = ev['member']

    # サークル別のコンテキスト
    circle_context = {
        '学びのAIくらし活用サークル': {
            'theme': 'AI技術の日常生活への活用',
            'target': 'AIに興味がある一般市民、主婦、シニア層',
            'goal': '最新のAI技術を身近な生活に取り入れ、暮らしの質を向上させる',
            'skills': 'AI活用スキル、デジタルリテラシー、創造的思考力',
        },
        '学びのデジタル安全生活サークル': {
            'theme': 'デジタル社会における安全・セキュリティ',
            'target': 'スマートフォンやインターネットを利用する全年齢層',
            'goal': 'デジタル社会で安全に暮らすための知識と実践力を身につける',
            'skills': '情報セキュリティ、フェイク情報の見分け方、プライバシー保護',
        },
        '学びの商店主・個人事業主サークル': {
            'theme': 'デジタル技術による事業支援・DX推進',
            'target': '地域の商店主、個人事業主、起業を考えている方',
            'goal': 'デジタルツールを活用して集客・業務効率化を実現する',
            'skills': 'DXスキル、マーケティング、業務効率化ツール活用',
        },
        '学びの写真メディア活用サークル': {
            'theme': '写真・映像メディアの活用と表現',
            'target': '写真や映像に興味がある方、SNS発信を学びたい方',
            'goal': 'デジタル写真・映像技術を学び、思い出の記録や発信力を高める',
            'skills': '撮影技術、画像編集、SNS活用、メディアリテラシー',
        },
        '学びのE-Sportsサークル': {
            'theme': 'ゲームを通じた学びと創造性の育成',
            'target': '小中学生、親子、ゲームを通じて学びたい方',
            'goal': 'ゲームを通じてプログラミング的思考、チームワーク、創造力を育む',
            'skills': 'プログラミング的思考、論理的思考力、チームワーク、創造力',
        },
    }

    ctx = circle_context.get(circle, {
        'theme': 'デジタル技術の学びと活用',
        'target': '地域住民',
        'goal': 'デジタル技術を学び、生活や仕事に活かす',
        'skills': 'デジタルスキル全般',
    })

    report = {
        'overview': f'本イベント「{title}」は、{circle}が主催する学びの場です。{ctx["theme"]}をテーマに、{venue}会場にて開催されます。参加者が実際に手を動かしながら学べる体験型のプログラムを提供し、地域のデジタルリテラシー向上に貢献します。',

        'purpose': f'本講座の目的は、{ctx["goal"]}ことです。「{title}」というテーマのもと、参加者一人ひとりが実践的なスキルを身につけ、日常生活や仕事の中で活用できるようになることを目指します。特に{ctx["target"]}を主な対象とし、わかりやすく丁寧な指導を行います。',

        'target_audience': ctx['target'],

        'content_detail': _generate_content_detail(title, circle, ctx),

        'schedule': f'''
【開催概要】
・開催日：{date}
・時間：{time}〜（約90分）
・会場：{venue}
・対象：{member if member else ctx["target"]}
・定員：20名（先着順）
・持ち物：スマートフォンまたはタブレット

【プログラム】
1. オープニング・自己紹介（10分）
2. テーマ解説・デモンストレーション（20分）
3. ハンズオン実践ワーク（40分）
4. 質疑応答・まとめ（15分）
5. アンケート記入・次回案内（5分）''',

        'expected_outcome': f'本講座を通じて、参加者は{ctx["skills"]}を習得できます。また、同じ関心を持つ地域の仲間とのつながりが生まれ、継続的な学びのコミュニティ形成にも寄与します。受講後は実生活ですぐに活用できる実践的な知識とスキルを持ち帰っていただけます。',

        'future': f'今後も{circle}では、月に複数回の定期的な講座を開催し、段階的にスキルアップできるカリキュラムを提供してまいります。参加者同士の交流会や、より高度な内容への発展講座も計画しています。地域全体のデジタル活用力向上を目指し、継続的な活動を展開します。',
    }
    return report


def _generate_content_detail(title, circle, ctx):
    """タイトルに基づいた講座内容の詳細を生成"""
    title_lower = title.lower()

    if 'AI' in title or 'ai' in title_lower or 'gemini' in title_lower or 'chatgpt' in title_lower:
        return f'''本講座「{title}」では、最新のAI技術を活用した実践的な内容をお届けします。

【講座のポイント】
・AIツールの基本的な使い方と活用のコツ
・実際の生活シーンを想定したハンズオン演習
・参加者一人ひとりのペースに合わせた丁寧なサポート
・すぐに使える実践テクニックの紹介

AIは難しいものではなく、私たちの生活を豊かにする身近なツールです。本講座では、初めての方でも安心して参加いただけるよう、基礎からわかりやすく解説します。実際にスマートフォンやタブレットを使いながら、楽しく学んでいきましょう。'''

    elif '詐欺' in title or 'セキュリティ' in title or '安全' in title or '守る' in title or 'フェイク' in title:
        return f'''本講座「{title}」では、デジタル社会で身を守るための実践的な知識をお伝えします。

【講座のポイント】
・最新のデジタル犯罪の手口と対策
・具体的な事例に基づく被害防止策
・すぐに実践できるセキュリティ設定
・家族や周囲の方への注意喚起のポイント

デジタル技術の進化に伴い、詐欺や不正アクセスの手口も巧妙化しています。本講座では、具体的な事例を交えながら、誰でもすぐに実践できる対策をわかりやすくお伝えします。ご自身だけでなく、ご家族の安全を守るための知識を身につけましょう。'''

    elif '商店' in circle or '事業主' in circle or 'DX' in title or '集客' in title or 'チラシ' in title:
        return f'''本講座「{title}」では、事業に直結するデジタル活用術を実践的に学びます。

【講座のポイント】
・すぐに業務に活かせるデジタルツールの使い方
・コストをかけずに始められる集客・PR手法
・実際の成功事例に基づくノウハウ共有
・参加者同士の情報交換と相互支援

地域の商店主・個人事業主の皆さまが、デジタルツールを活用して事業をより良くするための実践講座です。難しい専門知識は不要、今あるスマートフォンやパソコンですぐに始められる方法をお伝えします。'''

    elif '写真' in title or '撮影' in title or '動画' in title or 'カメラ' in title or 'Instagram' in title:
        return f'''本講座「{title}」では、写真や映像の撮影・編集・活用テクニックを実践的に学びます。

【講座のポイント】
・プロが教える撮影の基本テクニック
・スマートフォンでもできる本格的な撮影術
・簡単に使える編集ツールの活用法
・SNSでの効果的な発信方法

日常の何気ない瞬間を、より魅力的に残すための技術をお伝えします。特別な機材は不要、お手持ちのスマートフォンで参加いただけます。撮影のコツから編集、共有まで、一連の流れを体験しながら学びましょう。'''

    elif 'マイクラ' in title or 'マインクラフト' in title or 'フォートナイト' in title or 'Roblox' in title or '原神' in title or 'ゲーム' in title or 'スクラッチ' in title or 'プログラミング' in title:
        return f'''本講座「{title}」では、ゲームやプログラミングを通じた楽しい学びの場を提供します。

【講座のポイント】
・遊びながら自然に身につくプログラミング的思考
・チームワークとコミュニケーション力の育成
・創造力と問題解決能力の向上
・安全なゲーム環境での学習体験

ゲームは単なる娯楽ではなく、論理的思考やチームワーク、創造力を育む優れた教育ツールです。本講座では、楽しみながら自然とスキルが身につくプログラムを提供します。お子様だけでなく、親子での参加も大歓迎です。'''

    else:
        return f'''本講座「{title}」では、{ctx["theme"]}に関する実践的な内容をお届けします。

【講座のポイント】
・初心者にもわかりやすい丁寧な解説
・実際に手を動かして学べるハンズオン形式
・すぐに実生活で活用できる実践的な内容
・参加者同士の交流と学び合い

{ctx["target"]}を主な対象とし、楽しみながら学べるプログラムを用意しています。これまで参加したことがない方も、お気軽にご参加ください。'''


# ===== レポートHTML生成 =====
def generate_report_html(ev, report):
    month_names = {4:'April', 5:'May', 6:'June', 7:'July', 8:'August', 9:'September'}
    date_obj = datetime.strptime(ev['date'], '%Y-%m-%d')
    month_en = month_names.get(date_obj.month, '')

    html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{ev['title']} - イベントレポート</title>
<style>
@page {{
    size: A4;
    margin: 20mm;
}}
body {{
    font-family: "Hiragino Kaku Gothic ProN", "Meiryo", sans-serif;
    line-height: 1.8;
    color: #333;
    max-width: 210mm;
    margin: 0 auto;
    padding: 20mm;
    font-size: 10.5pt;
}}
.header {{
    border-bottom: 3px solid #2c5f8a;
    padding-bottom: 15px;
    margin-bottom: 20px;
}}
.header h1 {{
    font-size: 18pt;
    color: #2c5f8a;
    margin: 0 0 8px 0;
    line-height: 1.4;
}}
.meta-info {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 5px;
    font-size: 9.5pt;
    color: #555;
    background: #f5f8fb;
    padding: 10px 15px;
    border-radius: 5px;
}}
h2 {{
    font-size: 13pt;
    color: #2c5f8a;
    border-left: 4px solid #2c5f8a;
    padding-left: 10px;
    margin-top: 20px;
    margin-bottom: 10px;
}}
.section {{
    margin-bottom: 15px;
}}
.content-box {{
    background: #fafafa;
    border: 1px solid #e0e0e0;
    border-radius: 5px;
    padding: 12px 15px;
    margin: 10px 0;
    white-space: pre-line;
    font-size: 10pt;
}}
.footer {{
    margin-top: 30px;
    padding-top: 15px;
    border-top: 1px solid #ddd;
    font-size: 8.5pt;
    color: #888;
    text-align: center;
}}
.badge {{
    display: inline-block;
    background: #2c5f8a;
    color: white;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 9pt;
    margin-right: 5px;
}}
</style>
</head>
<body>
<div class="header">
    <h1>{ev['title']}</h1>
    <div class="meta-info">
        <div><strong>開催日：</strong>{ev['date']}</div>
        <div><strong>時間：</strong>{ev['time']}〜</div>
        <div><strong>会場：</strong>{ev['venue']}</div>
        <div><strong>主催：</strong>{ev['circle']}</div>
        <div><strong>対象：</strong>{ev['member'] if ev['member'] else '一般'}</div>
        <div><span class="badge">{ev['circle'].replace('学びの','').replace('サークル','')}</span></div>
    </div>
</div>

<div class="section">
    <h2>1. イベント概要</h2>
    <p>{report['overview']}</p>
</div>

<div class="section">
    <h2>2. 開催目的</h2>
    <p>{report['purpose']}</p>
</div>

<div class="section">
    <h2>3. 講座内容</h2>
    <div class="content-box">{report['content_detail']}</div>
</div>

<div class="section">
    <h2>4. 開催スケジュール</h2>
    <div class="content-box">{report['schedule']}</div>
</div>

<div class="section">
    <h2>5. 期待される成果</h2>
    <p>{report['expected_outcome']}</p>
</div>

<div class="section">
    <h2>6. 今後の展開</h2>
    <p>{report['future']}</p>
</div>

<div class="footer">
    <p>成田富里地域 {ev['circle']} | {ev['date']} {ev['venue']}会場</p>
    <p>本レポートは {ev['circle']} の活動記録として作成されました。</p>
</div>
</body>
</html>'''
    return html


# ===== プレゼンPDF生成 =====
def generate_presentation_pdf(ev, report, output_path):
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=20*mm,
        bottomMargin=20*mm,
        leftMargin=20*mm,
        rightMargin=20*mm,
    )

    styles = getSampleStyleSheet()

    # カスタムスタイル
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=FONT_GOTHIC,
        fontSize=22,
        leading=30,
        textColor=HexColor('#2c5f8a'),
        alignment=TA_CENTER,
        spaceAfter=15,
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName=FONT_GOTHIC,
        fontSize=14,
        leading=20,
        textColor=HexColor('#555555'),
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading1'],
        fontName=FONT_GOTHIC,
        fontSize=16,
        leading=22,
        textColor=HexColor('#2c5f8a'),
        spaceBefore=15,
        spaceAfter=10,
    )
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontName=FONT_MINCHO,
        fontSize=11,
        leading=18,
        textColor=HexColor('#333333'),
        alignment=TA_JUSTIFY,
        spaceAfter=8,
    )
    bullet_style = ParagraphStyle(
        'CustomBullet',
        parent=body_style,
        leftIndent=20,
        bulletIndent=10,
        spaceAfter=5,
    )
    small_style = ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontName=FONT_MINCHO,
        fontSize=9,
        leading=14,
        textColor=HexColor('#888888'),
        alignment=TA_CENTER,
    )
    slide_num_style = ParagraphStyle(
        'SlideNum',
        parent=styles['Normal'],
        fontName=FONT_GOTHIC,
        fontSize=9,
        textColor=HexColor('#aaaaaa'),
        alignment=TA_CENTER,
    )

    story = []

    # ===== スライド1: 表紙 =====
    story.append(Spacer(1, 60*mm))
    story.append(Paragraph(ev['title'], title_style))
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(ev['circle'], subtitle_style))
    story.append(Spacer(1, 5*mm))

    info_data = [[
        Paragraph(f"<b>{ev['date']}</b>", ParagraphStyle('info', parent=body_style, alignment=TA_CENTER, fontSize=12)),
        Paragraph(f"<b>{ev['time']}~</b>", ParagraphStyle('info', parent=body_style, alignment=TA_CENTER, fontSize=12)),
        Paragraph(f"<b>{ev['venue']}会場</b>", ParagraphStyle('info', parent=body_style, alignment=TA_CENTER, fontSize=12)),
    ]]
    info_table = Table(info_data, colWidths=[55*mm, 35*mm, 45*mm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), HexColor('#f0f5fa')),
        ('BOX', (0,0), (-1,-1), 1, HexColor('#2c5f8a')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, HexColor('#d0d8e0')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 30*mm))
    member_text = ev['member'] if ev['member'] else '一般'
    story.append(Paragraph(f"対象：{member_text}", subtitle_style))
    story.append(PageBreak())

    # ===== スライド2: 本日の内容 =====
    story.append(Paragraph("本日の内容", heading_style))
    story.append(Spacer(1, 5*mm))

    agenda_items = [
        "1. はじめに - 本講座の目的とゴール",
        "2. テーマ解説 - 基本知識の確認",
        "3. デモンストレーション - 実例紹介",
        "4. ハンズオン実践 - 実際にやってみよう",
        "5. まとめ・質疑応答",
    ]
    for item in agenda_items:
        story.append(Paragraph(item, bullet_style))
        story.append(Spacer(1, 3*mm))

    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(report['overview'], body_style))
    story.append(PageBreak())

    # ===== スライド3: 講座の目的 =====
    story.append(Paragraph("講座の目的", heading_style))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(report['purpose'], body_style))
    story.append(Spacer(1, 10*mm))

    story.append(Paragraph("対象者", heading_style))
    story.append(Paragraph(f"・{report['target_audience']}", bullet_style))
    story.append(PageBreak())

    # ===== スライド4: 講座内容詳細 =====
    story.append(Paragraph("講座内容", heading_style))
    story.append(Spacer(1, 5*mm))
    # content_detailを段落ごとに分割
    content_lines = report['content_detail'].split('\n')
    for line in content_lines:
        line = line.strip()
        if line:
            if line.startswith('【') or line.startswith('・'):
                story.append(Paragraph(line, ParagraphStyle(
                    'BoldLine', parent=body_style, fontName=FONT_GOTHIC, fontSize=11, leading=18,
                    spaceBefore=5, spaceAfter=3,
                )))
            else:
                story.append(Paragraph(line, body_style))
    story.append(PageBreak())

    # ===== スライド5: 実践ワーク =====
    story.append(Paragraph("実践ワーク", heading_style))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("実際に手を動かして体験してみましょう！", ParagraphStyle(
        'Emphasis', parent=body_style, fontName=FONT_GOTHIC, fontSize=13, textColor=HexColor('#2c5f8a'),
        alignment=TA_CENTER, spaceBefore=10, spaceAfter=15,
    )))

    work_items = _generate_work_items(ev['title'], ev['circle'])
    for i, item in enumerate(work_items, 1):
        story.append(Paragraph(f"Step {i}: {item}", bullet_style))
        story.append(Spacer(1, 3*mm))

    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("※ わからないことがあれば、お気軽にスタッフへお声がけください。", small_style))
    story.append(PageBreak())

    # ===== スライド6: まとめ =====
    story.append(Paragraph("まとめ", heading_style))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(report['expected_outcome'], body_style))
    story.append(Spacer(1, 10*mm))

    story.append(Paragraph("今後の活動予定", heading_style))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(report['future'], body_style))
    story.append(Spacer(1, 15*mm))

    story.append(Paragraph("ご参加ありがとうございました！", ParagraphStyle(
        'Thanks', parent=title_style, fontSize=18, textColor=HexColor('#2c5f8a'),
    )))
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(f"{ev['circle']} | {ev['venue']}会場", small_style))

    # PDF出力
    doc.build(story)


def _generate_work_items(title, circle):
    """タイトルに応じた実践ワークのステップを生成"""
    title_lower = title.lower()

    if 'AI' in title or 'ai' in title_lower or 'gemini' in title_lower or 'chatgpt' in title_lower:
        return [
            "AIツールを起動して、基本画面を確認しましょう",
            "まずは簡単な質問をAIに投げかけてみましょう",
            "今日のテーマに沿った実践課題に挑戦しましょう",
            "応用テクニックを試してみましょう",
            "作成した成果物を共有して感想を伝え合いましょう",
        ]
    elif '詐欺' in title or 'セキュリティ' in title or '安全' in title or '守る' in title:
        return [
            "自分のスマートフォンのセキュリティ設定を確認しましょう",
            "実際の詐欺メール・メッセージの見分け方を練習しましょう",
            "パスワードの安全性をチェックしてみましょう",
            "セキュリティ設定を最適化しましょう",
            "家族への伝え方をグループで考えてみましょう",
        ]
    elif '写真' in title or '撮影' in title or 'カメラ' in title:
        return [
            "カメラの基本設定を確認しましょう",
            "構図を意識して撮影してみましょう",
            "編集ツールで写真を加工してみましょう",
            "ビフォー・アフターを比較してみましょう",
            "お気に入りの一枚を選んで共有しましょう",
        ]
    elif 'マイクラ' in title or 'フォートナイト' in title or 'Roblox' in title or '原神' in title or 'ゲーム' in title:
        return [
            "ゲームの基本操作を確認しましょう",
            "今日のミッション・目標を確認しましょう",
            "チームで協力して課題に取り組みましょう",
            "振り返り：何を学んだか考えてみましょう",
            "次回に向けた目標を立てましょう",
        ]
    else:
        return [
            "今日使うツール・アプリを準備しましょう",
            "基本的な操作方法を確認しましょう",
            "実践課題に挑戦してみましょう",
            "応用テクニックを試してみましょう",
            "成果を共有して感想を伝え合いましょう",
        ]


# ===== Webサイト（index.html）生成 =====
def generate_website(events):
    # 月ごとにグループ化
    months_data = {}
    month_names_ja = {4:'4月', 5:'5月', 6:'6月', 7:'7月', 8:'8月', 9:'9月'}

    for ev in events:
        date_obj = datetime.strptime(ev['date'], '%Y-%m-%d')
        month_key = date_obj.month
        if month_key not in months_data:
            months_data[month_key] = []
        months_data[month_key].append(ev)

    # サークル別の色
    circle_colors = {
        '学びのAIくらし活用サークル': '#4a90d9',
        '学びのデジタル安全生活サークル': '#e74c3c',
        '学びの商店主・個人事業主サークル': '#27ae60',
        '学びの写真メディア活用サークル': '#f39c12',
        '学びのE-Sportsサークル': '#9b59b6',
    }

    event_cards = ''
    for month_key in sorted(months_data.keys()):
        month_events = months_data[month_key]
        event_cards += f'<h2 class="month-header">{month_names_ja[month_key]}</h2>\n'
        event_cards += '<div class="events-grid">\n'
        for ev in month_events:
            color = circle_colors.get(ev['circle'], '#666')
            circle_short = ev['circle'].replace('学びの','').replace('サークル','')
            event_cards += f'''<div class="event-card" onclick="location.href='{ev['folder_name']}/report.html'">
    <div class="card-header" style="border-left: 4px solid {color};">
        <span class="card-badge" style="background:{color};">{circle_short}</span>
        <span class="card-venue">{ev['venue']}</span>
    </div>
    <h3 class="card-title">{ev['title']}</h3>
    <div class="card-meta">
        <span class="card-date">{ev['date']}</span>
        <span class="card-time">{ev['time']}〜</span>
        <span class="card-member">{ev['member'] if ev['member'] else '一般'}</span>
    </div>
    <div class="card-links">
        <a href="{ev['folder_name']}/report.html" class="link-report">レポート</a>
        <a href="{ev['folder_name']}/presentation.pdf" class="link-pres">プレゼン資料</a>
    </div>
</div>
'''
        event_cards += '</div>\n'

    html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>成田富里地域 2026年度 イベントスケジュール</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
    font-family: "Hiragino Kaku Gothic ProN", "Meiryo", "Noto Sans JP", sans-serif;
    background: #f0f2f5;
    color: #333;
    line-height: 1.6;
}}
.container {{
    max-width: 1200px;
    margin: 0 auto;
    padding: 20px;
}}
.hero {{
    background: linear-gradient(135deg, #2c5f8a 0%, #4a90d9 100%);
    color: white;
    padding: 40px 30px;
    border-radius: 12px;
    margin-bottom: 30px;
    text-align: center;
}}
.hero h1 {{
    font-size: 28px;
    margin-bottom: 10px;
}}
.hero p {{
    font-size: 16px;
    opacity: 0.9;
}}
.stats {{
    display: flex;
    justify-content: center;
    gap: 30px;
    margin-top: 20px;
}}
.stat-item {{
    text-align: center;
}}
.stat-num {{
    font-size: 32px;
    font-weight: bold;
}}
.stat-label {{
    font-size: 12px;
    opacity: 0.8;
}}
.filter-bar {{
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 20px;
    padding: 15px;
    background: white;
    border-radius: 8px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
}}
.filter-btn {{
    padding: 6px 14px;
    border: 2px solid #ddd;
    border-radius: 20px;
    background: white;
    cursor: pointer;
    font-size: 13px;
    transition: all 0.2s;
}}
.filter-btn:hover, .filter-btn.active {{
    border-color: #2c5f8a;
    background: #2c5f8a;
    color: white;
}}
.month-header {{
    font-size: 22px;
    color: #2c5f8a;
    margin: 25px 0 15px;
    padding-bottom: 8px;
    border-bottom: 2px solid #2c5f8a;
}}
.events-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
    gap: 16px;
    margin-bottom: 20px;
}}
.event-card {{
    background: white;
    border-radius: 10px;
    padding: 18px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.08);
    cursor: pointer;
    transition: transform 0.2s, box-shadow 0.2s;
}}
.event-card:hover {{
    transform: translateY(-3px);
    box-shadow: 0 6px 16px rgba(0,0,0,0.12);
}}
.card-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding-left: 10px;
    margin-bottom: 10px;
}}
.card-badge {{
    color: white;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: bold;
}}
.card-venue {{
    font-size: 12px;
    color: #888;
    font-weight: bold;
}}
.card-title {{
    font-size: 15px;
    line-height: 1.5;
    margin-bottom: 10px;
    color: #222;
}}
.card-meta {{
    display: flex;
    gap: 12px;
    font-size: 12px;
    color: #777;
    margin-bottom: 12px;
}}
.card-links {{
    display: flex;
    gap: 8px;
}}
.card-links a {{
    display: inline-block;
    padding: 5px 14px;
    border-radius: 5px;
    font-size: 12px;
    text-decoration: none;
    font-weight: bold;
    transition: opacity 0.2s;
}}
.card-links a:hover {{ opacity: 0.8; }}
.link-report {{
    background: #e8f0fe;
    color: #2c5f8a;
}}
.link-pres {{
    background: #fef3e0;
    color: #e67e22;
}}
.footer {{
    text-align: center;
    padding: 30px;
    color: #999;
    font-size: 13px;
}}

/* フィルター機能 */
.event-card[data-circle] {{ display: block; }}
.hidden {{ display: none !important; }}

/* レスポンシブ */
@media (max-width: 768px) {{
    .events-grid {{ grid-template-columns: 1fr; }}
    .hero h1 {{ font-size: 22px; }}
    .stats {{ gap: 15px; }}
}}
</style>
</head>
<body>
<div class="container">
    <div class="hero">
        <h1>成田富里地域 2026年度<br>イベントスケジュール</h1>
        <p>学びのサークル活動 年間イベント一覧</p>
        <div class="stats">
            <div class="stat-item">
                <div class="stat-num">96</div>
                <div class="stat-label">イベント総数</div>
            </div>
            <div class="stat-item">
                <div class="stat-num">6</div>
                <div class="stat-label">ヶ月間</div>
            </div>
            <div class="stat-item">
                <div class="stat-num">3</div>
                <div class="stat-label">会場</div>
            </div>
            <div class="stat-item">
                <div class="stat-num">5</div>
                <div class="stat-label">サークル</div>
            </div>
        </div>
    </div>

    <div class="filter-bar">
        <button class="filter-btn active" onclick="filterEvents('all')">すべて</button>
        <button class="filter-btn" onclick="filterEvents('AIくらし活用')" style="border-color:#4a90d9;">AIくらし活用</button>
        <button class="filter-btn" onclick="filterEvents('デジタル安全生活')" style="border-color:#e74c3c;">デジタル安全生活</button>
        <button class="filter-btn" onclick="filterEvents('商店主・個人事業主')" style="border-color:#27ae60;">商店主・個人事業主</button>
        <button class="filter-btn" onclick="filterEvents('写真メディア活用')" style="border-color:#f39c12;">写真メディア活用</button>
        <button class="filter-btn" onclick="filterEvents('E-Sports')" style="border-color:#9b59b6;">E-Sports</button>
        <button class="filter-btn" onclick="filterEvents('富里')">📍富里</button>
        <button class="filter-btn" onclick="filterEvents('成田')">📍成田</button>
        <button class="filter-btn" onclick="filterEvents('鎌ヶ谷')">📍鎌ヶ谷</button>
    </div>

    {event_cards}

    <div class="footer">
        <p>成田富里地域 学びのサークル活動 2026年度</p>
    </div>
</div>

<script>
function filterEvents(keyword) {{
    const cards = document.querySelectorAll('.event-card');
    const btns = document.querySelectorAll('.filter-btn');
    btns.forEach(b => b.classList.remove('active'));
    event.target.classList.add('active');

    cards.forEach(card => {{
        if (keyword === 'all') {{
            card.classList.remove('hidden');
        }} else {{
            const text = card.textContent;
            if (text.includes(keyword)) {{
                card.classList.remove('hidden');
            }} else {{
                card.classList.add('hidden');
            }}
        }}
    }});
}}
</script>
</body>
</html>'''
    return html


# ===== メイン処理 =====
def main():
    print("イベントデータを読み込み中...")
    events = load_events()
    print(f"  {len(events)}件のイベントを読み込みました。")

    output_dir = os.path.join(BASE_DIR, '2026年度イベント')
    os.makedirs(output_dir, exist_ok=True)

    for i, ev in enumerate(events, 1):
        folder_path = os.path.join(output_dir, ev['folder_name'])
        os.makedirs(folder_path, exist_ok=True)

        # レポート生成
        report = generate_report_content(ev)
        report_html = generate_report_html(ev, report)
        with open(os.path.join(folder_path, 'report.html'), 'w', encoding='utf-8') as f:
            f.write(report_html)

        # プレゼンPDF生成
        pdf_path = os.path.join(folder_path, 'presentation.pdf')
        try:
            generate_presentation_pdf(ev, report, pdf_path)
        except Exception as e:
            print(f"  ⚠ PDF生成エラー [{ev['title'][:30]}]: {e}")

        print(f"  [{i:02d}/{len(events)}] {ev['folder_name']}")

    # Webサイト生成
    print("\nWebサイトを生成中...")
    website_html = generate_website(events)
    with open(os.path.join(output_dir, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(website_html)

    print(f"\n✅ 完了！")
    print(f"  出力先: {output_dir}")
    print(f"  フォルダ数: {len(events)}")
    print(f"  Webサイト: {os.path.join(output_dir, 'index.html')}")


if __name__ == '__main__':
    main()
